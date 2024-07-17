from enum import Enum
from itertools import product
from sklearn.metrics import mean_squared_error, mean_absolute_error, mean_absolute_percentage_error
from sklearn.model_selection import RandomizedSearchCV, TimeSeriesSplit
from sklearn.ensemble import RandomForestRegressor, ExtraTreesRegressor
import openpyxl
from datetime import datetime
import pandas as pd
import numpy as np
import os


class Target(Enum):
    BA11 = 0
    BA47 = 1
    full = 2

class Split(Enum):
    S60 = 0
    S70 = 1
    S80 = 2

class Normalize_Method(Enum):
    Log = 0
    MM = 1
    NN = 2

class DR_Method(Enum):
    PCA = 0
    ICA = 1
    KPCA = 2
    Isomap = 3

class Variance(Enum):
    V90 = 0
    V95 = 1

class WindowSize(Enum):
    W1 = 0
    W2 = 1
    W3 = 2

# Mapping from Enums to actual values
folder_map = {Target.BA11: 'BA11', Target.BA47: 'BA47', Target.full: 'full_data'}
target_map = {Target.BA11: 'BA11', Target.BA47: 'BA47', Target.full: 'full'}
split_map = {Split.S60: '60', Split.S70: '70', Split.S80: '80'}
n_method_map = {Normalize_Method.Log: 'log', Normalize_Method.MM: 'MM', Normalize_Method.NN: 'nonnormalized'}
dr_method_map = {DR_Method.PCA: 'PCA', DR_Method.ICA: 'ICA', DR_Method.KPCA: 'KPCA', DR_Method.Isomap: 'Isomap'}
variance_map = {Variance.V90: '90', Variance.V95: '95'}
excel_method_to_file = {'Log':'log', 'MinMax':'MM'}
window_size_map = { WindowSize.W1 : 'window1', WindowSize.W2 : 'window2', WindowSize.W3 : 'window3' }
sliding_window_sizes = [WindowSize.W1, WindowSize.W2, WindowSize.W3]
RANDOM_STATE = 42
# get the path to data
script_dir = os.path.dirname(__file__)
reduced_data_dir = os.path.join(script_dir, '..', 'data', 'reduced_data')
reduced_data_dir = os.path.normpath(reduced_data_dir)

encoded_data_dir = os.path.join(script_dir, '..', 'data', 'reduced_encoded')
encoded_data_dir = os.path.normpath(encoded_data_dir)

CNN_data_dir = os.path.join(script_dir, '..', 'data', 'reduced_CNN')
CNN_data_dir = os.path.normpath(CNN_data_dir)

CNN_flatten_data_dir = os.path.join(script_dir, '..', 'data', 'reduced_CNN_flatten')
CNN_flatten_data_dir = os.path.normpath(CNN_flatten_data_dir)

def filter_combinations(targets=None, splits=None, n_methods=None, DR_methods=None, variances=None, windows = None):
    """
    Generate combinations of parameters for filtering datasets.

    Parameters:
    targets (list of Target, optional): List of Target enums to include in the combinations. 
                                         Defaults to excluding 'full'.
    splits (list of Split, optional): List of Split enums to include in the combinations. 
                                      Defaults to including all splits.
    n_methods (list of Normalize_Method, optional): List of Normalize_Method enums to include in the combinations. 
                                                    Defaults to excluding 'nonnormalized'.
    DR_methods (list of DR_Method, optional): List of DR_Method enums to include in the combinations. 
                                              Defaults to including all DR methods.
    variances (list of Variance, optional): List of Variance enums to include in the combinations. 
                                            Defaults to including all variances.

    Returns:
    list of tuples: A list of tuples, each containing a unique combination of the specified parameters.
                    Each tuple contains (target, split, n_method, DR_method, variance).

    Example:
    combinations = filter_combinations(
        targets=[Target.BA11, Target.BA47],
        splits=[Split.S60, Split.S70],
        n_methods=[Normalize_Method.Log, Normalize_Method.MM],
        DR_methods=[DR_Method.PCA],
        variances=[Variance.V90]
    )
    """
    # Define default lists excluding 'full' and 'nonnormalized'
    default_targets = [Target.BA11, Target.BA47]
    default_n_methods = [Normalize_Method.Log, Normalize_Method.MM]
    
    # Use provided arrays or defaults
    effective_targets = targets if targets else default_targets
    effective_splits = splits if splits else list(Split)
    effective_n_methods = n_methods if n_methods else default_n_methods
    effective_DR_methods = DR_methods if DR_methods else []
    effective_variances = variances if variances else []
    effective_windows = windows if windows else []
    
    # Generate all combinations
    # if no dr and variance, only make combination of target, split and method
    if DR_methods == None and variances == None and windows == None:
        all_combinations = product(effective_targets, effective_splits, effective_n_methods)
        return list(all_combinations)
    
    elif windows == None:
        all_combinations = product(effective_targets, effective_splits, effective_n_methods, effective_DR_methods, effective_variances)
        # Filter combinations to exclude variance 95 with Isomap
        filtered_combinations = [
            combo for combo in all_combinations 
            if not (combo[4] == Variance.V95 and combo[3] == DR_Method.Isomap)
        ]
        # Return the filtered combinations
        return list(filtered_combinations)
    
    else: 
        all_combinations = product(effective_targets, effective_splits, effective_n_methods, effective_DR_methods, effective_variances, effective_windows)
        filtered_combinations = [
            combo for combo in all_combinations 
            if not (combo[4] == Variance.V95 and combo[3] == DR_Method.Isomap)
        ]
        return list(filtered_combinations)


# read from train_test_split_data
def read_data_file(target, split, n_method, split_xy = True):
    data_dir = os.path.join(script_dir, '..', 'data', 'train_test_split_data')
    data_dir = os.path.normpath(data_dir)

    # Construct the file paths for train
    train_name = f"{target_map[target]}_{split_map[split]}_{n_method_map[n_method]}_train.csv"
    test_name = f"{target_map[target]}_{split_map[split]}_{n_method_map[n_method]}_test.csv"

    train_file = os.path.join(data_dir, folder_map[target], train_name)
    test_file = os.path.join(data_dir, folder_map[target], test_name)
    
    # Load data
    train_data = pd.read_csv(train_file)
    test_data = pd.read_csv(test_file)

    if not split_xy:
        return train_data, test_data
    
    y_train = train_data.pop('TOD')
    X_train = train_data

    y_test = test_data.pop('TOD')
    X_test = test_data

    return X_train, y_train, X_test, y_test
    
# read from reduced data folder
def read_reduced_file(target, split, n_method, dr_method, variance):
    '''
    Input:
        target_data(BA11, BA47, Combine)
        normalize_method(log, MM, NN)
        split(60, 70, 80)
        DR_method(ICA, KPCA, PCA, Isomap)
        variance(90, 95)

    Output:
        X_train, y_train, X_test, y_test
    '''

    # Construct the file paths for train
    train_name = f"{target_map[target]}_{split_map[split]}_{n_method_map[n_method]}_{dr_method_map[dr_method]}_{variance_map[variance]}_train.csv"
    test_name = f"{target_map[target]}_{split_map[split]}_{n_method_map[n_method]}_{dr_method_map[dr_method]}_{variance_map[variance]}_test.csv"

    train_file = os.path.join(reduced_data_dir, train_name)
    test_file = os.path.join(reduced_data_dir, test_name)
    
    # Load data
    train_data = pd.read_csv(train_file)
    test_data = pd.read_csv(test_file)

    y_train = train_data.pop('TOD')
    X_train = train_data

    y_test = test_data.pop('TOD')
    X_test = test_data

    return X_train, y_train, X_test, y_test

# read from reduced data folder
def read_reduced_encoded_file(target, split, n_method, dr_method, variance, window_size):
    '''
    Input:
        target_data(BA11, BA47, Combine)
        normalize_method(log, MM, NN)
        split(60, 70, 80)
        DR_method(ICA, KPCA, PCA, Isomap)
        variance(90, 95)

    Output:
        X_train, y_train, X_test, y_test
    '''

    # TODO: change these to take dynamic window size, since this function will be passed as a parameter into another function, consider change train_test_model to 
    # take argument list
    train_name = f"{target_map[target]}_{split_map[split]}_{n_method_map[n_method]}_{window_size_map[window_size]}_{dr_method_map[dr_method]}_{variance_map[variance]}_train.csv"
    test_name = f"{target_map[target]}_{split_map[split]}_{n_method_map[n_method]}_{window_size_map[window_size]}_{dr_method_map[dr_method]}_{variance_map[variance]}_test.csv"

    train_file = os.path.join(encoded_data_dir, train_name)
    test_file = os.path.join(encoded_data_dir, test_name)
    
    # Load data
    train_data = pd.read_csv(train_file)
    test_data = pd.read_csv(test_file)

    y_train = train_data.pop('TOD')
    X_train = train_data

    y_test = test_data.pop('TOD')
    X_test = test_data

    return X_train, y_train, X_test, y_test

# read from CNN data folder
def read_reduced_CNN_file(target, split, n_method, dr_method, variance, window_size, flatten):
    '''
    Input:
        target_data(BA11, BA47, Combine)
        normalize_method(log, MM, NN)
        split(60, 70, 80)
        DR_method(ICA, KPCA, PCA, Isomap)
        variance(90, 95)

    Output:
        X_train, y_train, X_test, y_test
    '''

    train_name = f"{target_map[target]}_{split_map[split]}_{n_method_map[n_method]}_{window_size_map[window_size]}_{dr_method_map[dr_method]}_{variance_map[variance]}_train.csv"
    test_name = f"{target_map[target]}_{split_map[split]}_{n_method_map[n_method]}_{window_size_map[window_size]}_{dr_method_map[dr_method]}_{variance_map[variance]}_test.csv"


    if flatten:
        train_file = os.path.join(CNN_flatten_data_dir, train_name)
        test_file = os.path.join(CNN_flatten_data_dir, test_name)
    else:
        train_file = os.path.join(CNN_data_dir, train_name)
        test_file = os.path.join(CNN_data_dir, test_name)
    
    # Load data
    train_data = pd.read_csv(train_file)
    test_data = pd.read_csv(test_file)

    y_train = train_data.pop('TOD')
    X_train = train_data

    y_test = test_data.pop('TOD')
    X_test = test_data

    return X_train, y_train, X_test, y_test



def train_test_model(models, param_grids, combinations, n_iter=10, cv=5, random_state=42, verbose = False, save_result = False, use_numpy = False, data_read_function = read_reduced_file, data_process_function = None, windows = False, flatten=False):
    '''
    train_test_model
    Input:
        array of models
        array of model parameters
        combinations of files
        save_result: if save the result into a csv
        use_numpy: convert df to numpy format
        data_read_function: function to read data
        data_process_function: some data needs to be transform into other form before
                                feed into the model
        

    Output:
        result, write to csv.
    '''
    results = []
    
    for combination in combinations:
        
        # TODO: after change the data read function, change this place as well. should be taking a argument list
        if windows and flatten:
            target, split, n_method, dr_method, variance, window_size = combination
            X_train, y_train, X_test, y_test = data_read_function(target=target, n_method=n_method, split=split, dr_method=dr_method, variance=variance, window_size = window_size, flatten = flatten)
            if verbose:
                print(f"Processing: {target_map[target]}_{split_map[split]}_{n_method_map[n_method]}_{window_size_map[window_size]}_{dr_method_map[dr_method]}_{variance_map[variance]}")
        elif windows:
            target, split, n_method, dr_method, variance, window_size = combination
            X_train, y_train, X_test, y_test = data_read_function(target=target, n_method=n_method, split=split, dr_method=dr_method, variance=variance, window_size = window_size)
            if verbose:
                print(f"Processing: {target_map[target]}_{split_map[split]}_{n_method_map[n_method]}_{window_size_map[window_size]}_{dr_method_map[dr_method]}_{variance_map[variance]}")
        else: 
            target, split, n_method, dr_method, variance = combination
            X_train, y_train, X_test, y_test = data_read_function(target=target, n_method=n_method, split=split, dr_method=dr_method, variance=variance)
            if verbose:
                print(f"Processing: {target_map[target]}_{split_map[split]}_{n_method_map[n_method]}_{dr_method_map[dr_method]}_{variance_map[variance]}")
        

        if use_numpy:
            X_train = X_train.values.astype(np.float32)
            X_test = X_test.values.astype(np.float32)
            y_train = y_train.values.astype(np.float32)
            y_test = y_test.values.astype(np.float32)

            if data_process_function:
                X_train, y_train = data_process_function(X_train, y_train)
                X_test, y_test = data_process_function(X_test, y_test)

            

        for model_name, model in models.items():
            param_grid = param_grids[model_name]
            
            if verbose:
                print(f"Running {model_name} with parameters: {param_grid}")
            # Initialize RandomizedSearchCV
            tscv = TimeSeriesSplit(n_splits=cv)
            randomized_search = RandomizedSearchCV(model, param_distributions=param_grid, n_iter=n_iter, cv=tscv, random_state=random_state)
            try:
                randomized_search.fit(X_train, y_train)
                best_model = randomized_search.best_estimator_
                # Predict on test set
                y_pred = best_model.predict(X_test)
            except ValueError as e:
                print(f"Fit Failed, should only happen with CNN")
                continue
            
            
            
            
            # Calculate evaluation metrics
            mse = mean_squared_error(y_test, y_pred)
            mae = mean_absolute_error(y_test, y_pred)
            mape = mean_absolute_percentage_error(y_test, y_pred)
            rmse = mse ** 0.5
            smape = 100 * (2 * np.abs(y_test - y_pred) / (np.abs(y_test) + np.abs(y_pred))).mean()
            
            result = {
                'target': target_map[target],
                'split': split_map[split],
                'n_method': n_method_map[n_method],
                'dr_method': dr_method_map[dr_method],
                'variance': variance_map[variance],
                'model': model_name,
                'parameter_search': 'Random',
                'best_params': randomized_search.best_params_,
                'mse': mse,
                'mae': mae,
                'mape': mape,
                'rmse': rmse,
                'smape': smape
            }
            results.append(result)

            if verbose:
                print(f"Best parameters for {model_name}: {randomized_search.best_params_}")
                print(f"Metrics - MSE: {mse}, MAE: {mae}, MAPE: {mape}, RMSE: {rmse}, SMAPE: {smape}")

    if save_result:
        results_df = pd.DataFrame(results)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_dir = os.path.join(reduced_data_dir, '..', 'program_output')
        output_dir = os.path.normpath(output_dir)
        os.makedirs(output_dir, exist_ok=True)
        results_file = os.path.join(output_dir, f'model_results_{timestamp}.csv')
        results_df.to_csv(results_file, index=False)
    
    if verbose and save_result:
        print(f"Results saved to {results_file}")

    return results_df


def read_setup_info(sheet):
    '''
    Read info of this sheet
    '''
    setup_info = {
        "Train:Test Ratio": None,
        "Dimension Reduction Techniques": None,
        "Data Normalization Method": None,
        "Variance1": None,
        "Variance2": None
    }
    for row in sheet.iter_rows(min_row=1, max_row=3, min_col=1, max_col=2):
        key = row[0].value
        value = row[1].value
        if key in setup_info:
            if key == "Train:Test Ratio" and value:
                setup_info[key] = value.split(':')[0].strip()  # Extract the first number
            else:
                setup_info[key] = value
    # Read variance numbers from A5 and I5
    setup_info["Variance1"] = '90'
    setup_info["Variance2"] = '95'
    return setup_info

def get_model_row_mapping(sheet, start_col, verbose = False):
    model_row_mapping = {}
    for row in sheet.iter_rows(min_row=6, max_row=sheet.max_row, min_col=start_col, max_col=start_col):
        model_name = row[0].value
        if isinstance(model_name, str):
            standardized_model_name = model_name.strip().lower()
            if standardized_model_name not in model_row_mapping:
                model_row_mapping[standardized_model_name] = row[0].row
                if verbose:
                        print(f"Mapping model '{model_name}' to row {row[0].row}")
                elif verbose:
                    print(f"Model '{model_name}' already mapped to row {model_row_mapping[standardized_model_name]}")
    return model_row_mapping


# TODO: option 2 have 3 window sizes. either pass in folder path like 'performance_sheets\window2' 
# or pass the window size as a new parameter
def write_results_to_excel(results_df, target_folder = 'performance_sheets', windows = False, verbose=False):
    results_df['split'] = results_df['split'].astype(str)
    results_df['variance'] = results_df['variance'].astype(str)

    # Group results by target
    grouped_results = results_df.groupby('target')

    changes_made = False  # Flag to track if any changes are made

    for target_name, target_results in grouped_results:
        path = os.path.join(reduced_data_dir, '..', target_folder, f'{target_name} Overall Model Peformance Results.xlsx')
       
        # Load the existing Excel
        workbook = openpyxl.load_workbook(path)
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Read setup information from the sheet
            setup_info = read_setup_info(sheet)
            train_test_ratio = setup_info["Train:Test Ratio"]
            dr_technique = setup_info["Dimension Reduction Techniques"]
            normalization_method = setup_info["Data Normalization Method"]
            variance1 = setup_info["Variance1"]
            variance2 = setup_info["Variance2"]

            # Get model row mapping for each section
            model_row_mapping = get_model_row_mapping(sheet, start_col=1, verbose=False)
            
            # Determine matching results for variance1
            matching_results_var1 = target_results[
                (target_results['split'] == train_test_ratio) &
                (target_results['dr_method'] == dr_technique) &
                (target_results['n_method'] == excel_method_to_file[normalization_method]) &
                (target_results['variance'] == variance1)
            ]
            
            # Determine matching results for variance2
            matching_results_var2 = target_results[
                (target_results['split'] == train_test_ratio) &
                (target_results['dr_method'] == dr_technique) &
                (target_results['n_method'] == excel_method_to_file[normalization_method]) &
                (target_results['variance'] == variance2)
            ]
            
            if matching_results_var1.empty and matching_results_var2.empty:
                continue
            # Base column numbers for variance sections
            base_col_var1 = 2
            base_col_var2 = 10

            # Write results for variance1
            for _, result in matching_results_var1.iterrows():
                model_name = result['model'].strip().lower()
                if model_name in model_row_mapping:
                    row_num = model_row_mapping[model_name]
                    sheet.cell(row=row_num, column=base_col_var1, value=result['parameter_search'])
                    sheet.cell(row=row_num, column=base_col_var1 + 1, value=str(result['best_params']))
                    sheet.cell(row=row_num, column=base_col_var1 + 2, value=result['mse'])
                    sheet.cell(row=row_num, column=base_col_var1 + 3, value=result['mae'])
                    sheet.cell(row=row_num, column=base_col_var1 + 4, value=result['mape'])
                    sheet.cell(row=row_num, column=base_col_var1 + 5, value=result['rmse'])
                    sheet.cell(row=row_num, column=base_col_var1 + 6, value=result['smape'])
                    changes_made = True
            
            # Write results for variance2
            for _, result in matching_results_var2.iterrows():
                model_name = result['model'].strip().lower()
                if model_name in model_row_mapping:
                    row_num = model_row_mapping[model_name]
                    sheet.cell(row=row_num, column=base_col_var2, value=result['parameter_search'])
                    sheet.cell(row=row_num, column=base_col_var2 + 1, value=str(result['best_params']))
                    sheet.cell(row=row_num, column=base_col_var2 + 2, value=result['mse'])
                    sheet.cell(row=row_num, column=base_col_var2 + 3, value=result['mae'])
                    sheet.cell(row=row_num, column=base_col_var2 + 4, value=result['mape'])
                    sheet.cell(row=row_num, column=base_col_var2 + 5, value=result['rmse'])
                    sheet.cell(row=row_num, column=base_col_var2 + 6, value=result['smape'])
                    changes_made = True
        
        # Save the workbook to the specified output path
        if changes_made:
            workbook.save(path)
            if verbose:
                print(f"Results saved to {path}")
        
        

# Below is an example of how to use previous functions
if __name__ == "__main__":

    # Define models
    models = {
        'Random Forest Regressor': RandomForestRegressor(random_state=42),
        'ExtraTreesRegressor': ExtraTreesRegressor(random_state=42)
    }

    # Define parameter grids for RandomizedSearchCV
    param_grids = {
        'Random Forest Regressor': {
            'n_estimators': [10, 50, 100, 200],
            'max_depth': [None, 10, 20, 30],
            'min_samples_split': [2, 5, 10],
            'min_samples_leaf': [1, 2, 4]
        },
        'ExtraTreesRegressor': {
            'n_estimators': [10, 50, 100, 200],
            'max_depth': [None, 10, 20, 30],
            'min_samples_split': [2, 5, 10],
            'min_samples_leaf': [1, 2, 4]
        }
    }

    # Specify which datasets to use
    combinations = filter_combinations(
        targets=[Target.BA11],
        splits=[Split.S60],
        n_methods=[Normalize_Method.Log],
        DR_methods=[DR_Method.ICA],
        variances=[Variance.V90]
    )
    
    results_df = train_test_model(models, param_grids, combinations, verbose=True, save_result=True)
    print(results_df)
    write_results_to_excel(results_df, verbose=True)
